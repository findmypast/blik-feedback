import csv
import io
import re
from collections import Counter, defaultdict

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import (
    OrganizationInvitation,
    OrganizationRole,
    Reviewee,
    Team,
    TeamLeadGrant,
    UserProfile,
)
from accounts.name_utils import normalize_name_part
from accounts.permissions import assign_organization_admin, assign_organization_member

MAX_ROWS = 500
MAX_FILE_SIZE = 5 * 1024 * 1024
REQUIRED_FIELDS = ('first_name', 'last_name', 'email')
OPTIONAL_FIELDS = ('team', 'manager', 'role')


class PeopleImportError(ValidationError):
    pass


def _clean(value):
    return str(value or '').strip()


def _team_names(value):
    """Return unique, trimmed team names from a comma-separated cell."""
    names = []
    seen = set()
    for item in _clean(value).split(','):
        name = item.strip()
        key = name.casefold()
        if name and key not in seen:
            names.append(name)
            seen.add(key)
    return names


def _name_from_email(email):
    local_part = email.partition('@')[0]
    words = re.sub(r'[._+-]+', ' ', local_part)
    return normalize_name_part(words)


def read_people_file(upload):
    if upload.size > MAX_FILE_SIZE:
        raise PeopleImportError('The file must be 5 MB or smaller.')
    suffix = upload.name.lower().rsplit('.', 1)[-1]
    try:
        if suffix == 'csv':
            text = upload.read().decode('utf-8-sig')
            rows = list(csv.reader(io.StringIO(text)))
        elif suffix == 'xlsx':
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            workbook.close()
        else:
            raise PeopleImportError('Upload a CSV or XLSX file.')
    except (UnicodeDecodeError, csv.Error, OSError, ValueError) as exc:
        raise PeopleImportError('The spreadsheet could not be read.') from exc
    while rows and not any(_clean(value) for value in rows[-1]):
        rows.pop()
    if not rows or not any(_clean(value) for value in rows[0]):
        raise PeopleImportError('The first row must contain column headers.')
    if len(rows) - 1 > MAX_ROWS:
        raise PeopleImportError(f'Imports are limited to {MAX_ROWS} people.')
    headers = [_clean(value) for value in rows[0]]
    if any(not header for header in headers):
        raise PeopleImportError('Every column in the first row must have a header.')
    folded = [header.casefold() for header in headers]
    if len(folded) != len(set(folded)):
        raise PeopleImportError('Column headers must be unique.')
    data = [row for row in rows[1:] if any(_clean(value) for value in row)]
    return headers, data


def _mapped_rows(headers, raw_rows, mapping):
    missing = [field for field in REQUIRED_FIELDS if not mapping.get(field)]
    if missing:
        raise PeopleImportError('Map First Name, Last Name, and Work Email.')
    selected = [value for value in mapping.values() if value]
    if len(selected) != len(set(selected)):
        raise PeopleImportError('A spreadsheet column can only be mapped once.')
    unknown = set(selected) - set(headers)
    if unknown:
        raise PeopleImportError('One or more mapped headers are no longer present.')
    indexes = {field: headers.index(header) for field, header in mapping.items() if header}
    result = []
    for number, raw in enumerate(raw_rows, start=2):
        result.append({
            'row': number,
            **{
                field: _clean(raw[index] if index < len(raw) else '')
                for field, index in indexes.items()
            },
        })
    return result


def validate_people_import(organization, upload, mapping, role_mapping, mode):
    if mode not in {'update', 'sync'}:
        raise PeopleImportError('Select a valid import mode.')
    headers, raw_rows = read_people_file(upload)
    rows = _mapped_rows(headers, raw_rows, mapping)
    errors = []
    seen_emails = {}
    existing_profiles = list(
        UserProfile.objects.for_organization(organization).select_related('user')
    )
    profiles_by_email = {p.user.email.casefold(): p for p in existing_profiles}
    unavailable_emails = {
        email.casefold()
        for email in get_user_model().objects.exclude(
            profile__organization=organization
        ).exclude(email='').values_list('email', flat=True)
    }
    invitations_by_email = {
        invite.email.casefold(): invite for invite in OrganizationInvitation.objects.filter(
            organization=organization, accepted_at__isnull=True
        )
    }
    roles = {
        str(role.pk): role for role in OrganizationRole.objects.for_organization(organization)
    }
    existing_teams = {
        team.name.casefold(): team for team in Team.objects.for_organization(organization)
    }

    names = defaultdict(list)
    for profile in existing_profiles:
        name = profile.user.get_full_name().strip()
        if name:
            names[name.casefold()].append(profile.user.email.casefold())
    for row in rows:
        email = row.get('email', '').casefold()
        row['first_name'] = normalize_name_part(row.get('first_name'))
        row['last_name'] = normalize_name_part(row.get('last_name'))
        full_name = f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
        row['email'] = email
        row['name'] = full_name
        if not full_name:
            errors.append({'row': row['row'], 'message': 'Enter a first or last name.'})
        try:
            validate_email(email)
        except ValidationError:
            errors.append({'row': row['row'], 'message': 'Work Email is invalid.'})
        if email in seen_emails:
            errors.append({
                'row': row['row'],
                'message': f'Email duplicates row {seen_emails[email]}.',
            })
        seen_emails[email] = row['row']
        if email in unavailable_emails:
            errors.append({
                'row': row['row'],
                'message': 'This email cannot be imported into the organisation.',
            })
        if full_name:
            names[full_name.casefold()].append(email)

    manager_names_by_email = {
        profile.user.email.casefold(): (
            ' '.join(filter(None, (
                normalize_name_part(profile.user.first_name),
                normalize_name_part(profile.user.last_name),
            ))) or profile.user.email
        )
        for profile in existing_profiles
    }
    manager_names_by_email.update({
        row['email']: row['name'] for row in rows if row['name']
    })
    for row in rows:
        manager = row.get('manager', '').strip()
        manager_email = ''
        if manager:
            try:
                validate_email(manager)
                manager_email = manager.casefold()
                if manager_email in unavailable_emails:
                    errors.append({
                        'row': row['row'],
                        'message': 'This manager email cannot be imported into the organisation.',
                    })
            except ValidationError:
                matches = list(dict.fromkeys(names.get(manager.casefold(), [])))
                if len(matches) != 1:
                    message = 'Manager name is ambiguous.' if matches else 'Manager name was not found.'
                    errors.append({'row': row['row'], 'message': message})
                else:
                    manager_email = matches[0]
        row['manager_email'] = manager_email
        if manager_email:
            manager_names_by_email.setdefault(
                manager_email, _name_from_email(manager_email)
            )

    used_team_names = set(existing_teams)
    for row in rows:
        used_team_names.update(name.casefold() for name in _team_names(row.get('team')))
    untitled_number = 1
    for row in rows:
        team_names = _team_names(row.get('team'))
        if team_names:
            row['teams'] = team_names
            row['team'] = ', '.join(team_names)
            row['generated_team'] = False
            continue
        manager_name = manager_names_by_email.get(row.get('manager_email'))
        if manager_name:
            row['team'] = f"{manager_name}'s team"
            row['teams'] = [row['team']]
            row['generated_team'] = True
            used_team_names.add(row['team'].casefold())
            continue
        while f'untitled-{untitled_number}' in used_team_names:
            untitled_number += 1
        row['team'] = f'Untitled-{untitled_number}'
        row['teams'] = [row['team']]
        row['generated_team'] = True
        used_team_names.add(row['team'].casefold())
        untitled_number += 1

    distinct_roles = sorted({row.get('role', '') for row in rows if row.get('role')})
    for value in distinct_roles:
        mapped = role_mapping.get(value)
        if not mapped:
            errors.append({'row': None, 'message': f'Map the role “{value}”.'})
        elif (
            mapped.startswith('custom:')
            and mapped.removeprefix('custom:') not in roles
        ) or (
            mapped not in {'member', 'team_leader', 'admin'}
            and not mapped.startswith('custom:')
        ):
            errors.append({'row': None, 'message': f'Role mapping for “{value}” is invalid.'})

    for row in rows:
        row['action'] = (
            'Update existing person' if row['email'] in profiles_by_email
            else 'Update pending invitation' if row['email'] in invitations_by_email
            else 'Invite new person'
        )
        new_teams = [
            name for name in row['teams'] if name.casefold() not in existing_teams
        ]
        row['team_action'] = (
            'Create ' + ', '.join(new_teams) if new_teams else ''
        )
        row['mapped_role'] = role_mapping.get(row.get('role', ''), 'member')

    imported_emails = set(seen_emails)
    missing_profiles = []
    if mode == 'sync':
        missing_profiles = [
            p for p in existing_profiles
            if p.user.email.casefold() not in imported_emails
            and p.user.is_active
            and not p.user.is_superuser
            and not p.user.has_perm('accounts.can_manage_organization')
        ]
    manager_invitations = sorted({
        row['manager_email']
        for row in rows
        if row.get('manager_email')
        and row['manager_email'] not in profiles_by_email
        and row['manager_email'] not in imported_emails
    })
    return {
        'headers': headers,
        'rows': rows,
        'errors': errors,
        'roles': distinct_roles,
        'missing': [
            {'id': p.pk, 'name': p.user.get_full_name() or p.user.email, 'email': p.user.email}
            for p in missing_profiles
        ],
        'manager_invitations': manager_invitations,
        '_missing_profiles': missing_profiles,
    }


def apply_people_import(organization, preview, invited_by, deactivate_missing=False):
    invitations = []
    counts = defaultdict(int)
    with transaction.atomic():
        teams = {
            team.name.casefold(): team for team in Team.objects.for_organization(organization)
        }
        for row in preview['rows']:
            row_teams = []
            for team_name in row.get('teams', _team_names(row.get('team'))):
                team, created = Team.objects.get_or_create(
                    organization=organization,
                    name__iexact=team_name,
                    defaults={'name': team_name},
                )
                teams[team_name.casefold()] = team
                row_teams.append(team)
                counts['teams_created'] += int(created)
            team = row_teams[0] if row_teams else None

            profile = UserProfile.objects.for_organization(organization).filter(
                user__email__iexact=row['email']
            ).select_related('user').first()
            custom_role = None
            mapped_role = row['mapped_role']
            if mapped_role.startswith('custom:'):
                custom_role = OrganizationRole.objects.get(
                    organization=organization,
                    pk=mapped_role.removeprefix('custom:'),
                )
            if profile:
                user = profile.user
                user.first_name = row.get('first_name', '')[:150]
                user.last_name = row.get('last_name', '')[:150]
                user.is_active = True
                user.save(update_fields=['first_name', 'last_name', 'is_active'])
                if row.get('role'):
                    profile.organization_role = custom_role
                    profile.save(update_fields=['organization_role', 'updated_at'])
                if mapped_role == 'admin':
                    assign_organization_admin(user)
                elif row.get('role'):
                    assign_organization_member(user, profile.can_create_cycles_for_others)
                reviewee, _ = Reviewee.objects.get_or_create(
                    organization=organization,
                    email__iexact=row['email'],
                    defaults={'email': row['email'], 'name': row['name']},
                )
                reviewee.profile = profile
                reviewee.name = row['name']
                reviewee.team = team
                reviewee.is_active = True
                reviewee.save()
                retained_teams = list(profile.managed_teams.all())
                reviewee.teams.set(
                    list({item.pk: item for item in row_teams + retained_teams}.values())
                )
                if mapped_role == 'team_leader':
                    for assigned_team in row_teams:
                        TeamLeadGrant.objects.get_or_create(
                            profile=profile, team=assigned_team
                        )
                counts['updated'] += 1
            else:
                invitation = OrganizationInvitation.objects.filter(
                    organization=organization,
                    email__iexact=row['email'],
                ).first()
                created = invitation is None
                invitation = invitation or OrganizationInvitation(organization=organization)
                invitation.email = row['email']
                invitation.first_name = row.get('first_name', '')[:150]
                invitation.last_name = row.get('last_name', '')[:150]
                invitation.team = team
                invitation.invited_by = invited_by
                invitation.expires_at = timezone.now() + timezone.timedelta(days=7)
                invitation.accepted_at = None
                invitation.requested_role = (
                    mapped_role if not mapped_role.startswith('custom:') else 'member'
                )
                invitation.organization_role = custom_role
                invitation.save()
                reviewee, _ = Reviewee.objects.get_or_create(
                    organization=organization,
                    email__iexact=row['email'],
                    defaults={'email': row['email'], 'name': row['name']},
                )
                reviewee.name = row['name']
                reviewee.team = team
                reviewee.is_active = True
                reviewee.save()
                reviewee.teams.set(row_teams)
                invitations.append(invitation)
                counts['invited' if created else 'invitations_updated'] += 1

        profiles_by_email = {
            p.user.email.casefold(): p
            for p in UserProfile.objects.for_organization(organization).select_related('user')
        }
        manager_teams = defaultdict(list)
        team_reporting_managers = defaultdict(Counter)
        for row in preview['rows']:
            if row.get('manager_email'):
                manager_teams[row['manager_email']].extend(
                    teams[name.casefold()] for name in row.get('teams', [])
                )
                for name in row.get('teams', []):
                    team_reporting_managers[name.casefold()][row['manager_email']] += 1
        inferred_team_managers = {}
        for team_key, manager_counts in team_reporting_managers.items():
            manager_email, report_count = manager_counts.most_common(1)[0]
            runner_up_count = manager_counts.most_common(2)[1][1] \
                if len(manager_counts) > 1 else 0
            if report_count > runner_up_count:
                inferred_team_managers[team_key] = manager_email
        inferred_leader_emails = set(inferred_team_managers.values())

        invitation_emails = {invitation.email.casefold() for invitation in invitations}
        for manager_email, managed_teams in manager_teams.items():
            if manager_email in profiles_by_email:
                continue
            invitation = OrganizationInvitation.objects.filter(
                organization=organization,
                email__iexact=manager_email,
            ).first()
            created = invitation is None
            invitation = invitation or OrganizationInvitation(organization=organization)
            invitation.email = manager_email
            invitation.team = invitation.team or managed_teams[0]
            invitation.invited_by = invited_by
            invitation.expires_at = timezone.now() + timezone.timedelta(days=7)
            invitation.accepted_at = None
            if invitation.requested_role != 'admin':
                invitation.requested_role = (
                    'team_leader'
                    if manager_email in inferred_leader_emails else 'reporting_manager'
                )
            invitation.save()
            manager_reviewee, _ = Reviewee.objects.get_or_create(
                organization=organization,
                email__iexact=manager_email,
                defaults={'email': manager_email, 'name': manager_email},
            )
            if manager_reviewee.team_id is None:
                manager_reviewee.team = managed_teams[0]
                manager_reviewee.save(update_fields=['team', 'updated_at'])
            manager_reviewee.teams.add(*managed_teams)
            if manager_email not in invitation_emails:
                invitations.append(invitation)
                invitation_emails.add(manager_email)
                counts['invited' if created else 'invitations_updated'] += 1

        for team_key, manager_email in inferred_team_managers.items():
            team = teams[team_key]
            if team.manager_id:
                continue
            manager = profiles_by_email.get(manager_email)
            if manager:
                team.manager = manager
                team.pending_manager_email = ''
                team.save(update_fields=['manager', 'pending_manager_email', 'updated_at'])
                TeamLeadGrant.objects.get_or_create(profile=manager, team=team)
                manager_reviewee = Reviewee.objects.filter(
                    organization=organization, email__iexact=manager_email
                ).first()
                if manager_reviewee:
                    manager_reviewee.teams.add(team)
            else:
                team.pending_manager_email = manager_email
                team.save(update_fields=['pending_manager_email', 'updated_at'])
            counts['team_leaders_assigned'] += 1

        for row in preview['rows']:
            if not row.get('manager_email'):
                continue
            manager = profiles_by_email.get(row['manager_email'])
            employee = Reviewee.objects.get(
                organization=organization, email__iexact=row['email']
            )
            employee.reporting_manager = manager
            employee.pending_reporting_manager_email = '' if manager else row['manager_email']
            employee.save(update_fields=[
                'reporting_manager', 'pending_reporting_manager_email', 'updated_at'
            ])
            counts['managers_assigned'] += 1

        if deactivate_missing:
            for profile in preview['_missing_profiles']:
                profile.user.is_active = False
                profile.user.save(update_fields=['is_active'])
                Reviewee.objects.filter(profile=profile).update(is_active=False)
                counts['deactivated'] += 1
    return dict(counts), invitations
